import openpyxl as pyxl
import pandas as pd
import pyexcel

class Hidrologia(object):
    def __init__(self):
        pass

    def get_rdh(self, rdhs: list, coords: dict={'row':[8, 172], 'col':[1, 28]}):

        aux = []
        #  Itera sobre rdhs passados
        for k, rdh in enumerate(rdhs):

            # Abre arquivo
            wb = pyxl.load_workbook(
                filename=rdh,
                read_only=True,
            )
            ws = wb.worksheets[0]
            dt = pd.to_datetime(ws.cell(row=2, column=21).value[-10:], format='%d/%m/%Y')

            for i, row in enumerate(
                    ws.iter_rows(
                        min_row=coords['row'][0], max_row=coords['row'][1],
                        min_col=coords['col'][0], max_col=coords['col'][1]
                    )
            ):
                #  Testes de verificacao
                if row[4].data_type == 'n':

                    try:  # vazao natural
                        val_vaz_natr = float(row[13].value)
                    except(ValueError, TypeError):
                        val_vaz_natr = 0.0

                    try:  # cota
                        val_cota = float(row[14].value)
                    except(ValueError, TypeError):
                        val_cota = 0.0

                    try:  # arm
                        val_arm = float(row[15].value) / 100
                    except(ValueError, TypeError):
                        val_arm = 0.0

                    try:  # vazao vertida
                        val_vaz_vert = float(row[18].value)
                    except(ValueError, TypeError):
                        val_vaz_vert = 0.0

                    try:  # vazao defluida
                        val_vaz_defl = float(row[20].value)
                    except(ValueError, TypeError):
                        val_vaz_defl = 0.0

                    try:  # vazao afluente
                        val_vaz_aflu = float(row[22].value)
                    except(ValueError, TypeError):
                        val_vaz_aflu = 0.0

                    try:  # vazao incremental
                        val_vaz_inc = float(row[24].value)
                    except(ValueError, TypeError):
                        val_vaz_inc = 0.0

                    aux.append(
                        dict(
                            num_posto=row[4].value,
                            dat_medicao=dt,
                            val_vaz_natr=val_vaz_natr,
                            val_cota=val_cota,
                            val_arm=val_arm,
                            val_vaz_vert=val_vaz_vert,
                            val_vaz_defl=val_vaz_defl,
                            val_vaz_aflu=val_vaz_aflu,
                            val_vaz_inc=val_vaz_inc
                        )
                    )

            wb.close()

        df_vazao = pd.DataFrame(data=aux)
        self.dados = df_vazao
        return self.dados

    def get_acomph(self, acomphs, config_acomph: dict={'row':[6, 35], 'bloco_dados':9}):
        df_vazao = pd.DataFrame()

        # Itera sobre todos os acomphs passados
        for acomph in acomphs:

            try:
                wb = pyxl.load_workbook(
                    filename=acomph,
                    read_only=False,
                    data_only=True
                )
            except:
                pyexcel.save_book_as(
                    file_name=acomph,
                    dest_file_name='{}.xlsx'.format(acomph[:-4])
                )

                wb = pyxl.load_workbook(
                    filename='{}.xlsx'.format(acomph[:-4]),
                    read_only=False,
                    data_only=True
                )

            # Itera sobre todas as worksheets
            for sheet in wb.worksheets:
                dat_medicao = []
                for row in sheet.iter_cols(
                    min_row=config_acomph['row'][0], max_row=config_acomph['row'][1],
                    min_col=1, max_col=1
                ):
                    [dat_medicao.append(cell.value) for cell in row]

                # pega dados do posto
                for i in list(range(9, sheet.max_column - 1 + 9, config_acomph['bloco_dados'])):
                    vaz_natr = []
                    vaz_incr = []
                    vaz_aflu = []
                    vaz_defl = []
                    val_cota = []
                    num_posto = []

                    # vazao natural
                    for row in sheet.iter_cols(
                            min_row=config_acomph['row'][0], max_row=config_acomph['row'][1],
                            min_col=i, max_col=i
                    ):
                        [vaz_natr.append(float(cell.value)) for cell in row]

                    # vazao incremental
                    for row in sheet.iter_cols(
                            min_row=config_acomph['row'][0], max_row=config_acomph['row'][1],
                            min_col=i-1, max_col=i-1
                    ):
                        [vaz_incr.append(float(cell.value)) for cell in row]

                    # vazao afluente
                    for row in sheet.iter_cols(
                            min_row=config_acomph['row'][0], max_row=config_acomph['row'][1],
                            min_col=i-2, max_col=i-2
                    ):
                        [vaz_aflu.append(float(cell.value)) for cell in row]

                    # vazao defluente
                    for row in sheet.iter_cols(
                            min_row=config_acomph['row'][0], max_row=config_acomph['row'][1],
                            min_col=i-4, max_col=i-4
                    ):
                        [vaz_defl.append(float(cell.value)) for cell in row]

                    # vazao cota
                    for row in sheet.iter_cols(
                            min_row=config_acomph['row'][0], max_row=config_acomph['row'][1],
                            min_col=i-6, max_col=i-6
                    ):
                        [val_cota.append(float(cell.value)) for cell in row]

                    # numero do posto
                    [num_posto.append(int(sheet.cell(row=1, column=i).value)) for j in range(val_cota.__len__())]

                    df_aux = pd.DataFrame.from_dict(
                        dict(
                            num_posto=num_posto,
                            dat_medicao=dat_medicao,
                            val_vaz_natr=vaz_natr,
                            val_vaz_incr=vaz_incr,
                            val_vaz_aflu=vaz_aflu,
                            val_vaz_defl=vaz_defl,
                            val_cota=val_cota
                        )
                    )

                    if num_posto[0] != None:
                        df_vazao = pd.concat([df_vazao, df_aux])

        pass

        df_vazao.to_csv(
            path_or_buf='acomph.csv', sep=';',
            decimal=',', index=False,
            float_format='%5.2f',
            date_format='%Y-%m-%d'
        )
        self.dados = df_vazao
        return


class Calculador(object):

    def realiza_calculo(self, dados, posto):
        posto.calcula(dados)
