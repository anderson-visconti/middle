# !/usr/bin/env python
# *- coding: utf-8 -*-
import openpyxl
import os

configs = {
    'paths_excel': [
        r'C:\Onedrive\Middle Office\Middle\Decks\Prospectivos\2018\05\p1-pessimista\01\vazoes\01-LP.xlsm',
        r'C:\Onedrive\Middle Office\Middle\Decks\Prospectivos\2018\05\p1-pessimista\02\vazoes\02-LP.xlsm',
        r'C:\Onedrive\Middle Office\Middle\Decks\Prospectivos\2018\05\p1-pessimista\03\vazoes\03-LP.xlsm',
        r'C:\Onedrive\Middle Office\Middle\Decks\Prospectivos\2018\05\p1-pessimista\04\vazoes\04-LP.xlsm',
        r'C:\Onedrive\Middle Office\Middle\Decks\Prospectivos\2018\05\p1-pessimista\05\vazoes\05-LP.xlsm',
    ],

    'sheet_name': r'Resultado Final',
    'vazpast': r'VazPast',
    'coords_resultado': {
        'ini': {'row': 4, 'col': 3},
        'fim': {'row': 163, 'col': 14}
    },
    'coords_vazpast': {
        'ini': {'row': 6, 'col': 5},
        'fim': {'row': 221, 'col': 18}
    },

}

# Escrita dos prevs
for i, arquivo in enumerate(configs['paths_excel']):
    work_book = openpyxl.load_workbook(filename=arquivo, data_only=True)
    sheet = work_book[configs['sheet_name']]
    sheet_vazpast = work_book[configs['vazpast']]

    # Itera sobre o range completo
    for rows in sheet.iter_cols(
            min_row=configs['coords_resultado']['ini']['row'],
            min_col=configs['coords_resultado']['ini']['col'],
            max_row=configs['coords_resultado']['fim']['row'],
            max_col=configs['coords_resultado']['fim']['col']
    ):
        prevs = open(
            os.path.join(
                os.path.dirname(arquivo),
                'prevs_{}_{}.rv0'.format(
                    os.path.basename(os.path.splitext(arquivo)[0]),
                    sheet.cell(row=2, column=rows[0].col_idx).internal_value,
                )
            ),
            'w'
        )

        for row, posto in enumerate(rows):
            prevs.write(
                '{:6d}''{:5d}''{:10d}''{:10d}''{:10d}''{:10d}''{:10d}''{:10d}\n'.format(
                    row + 1,
                    sheet.cell(row=posto.row, column=2).value,
                    posto.internal_value,
                    posto.internal_value,
                    posto.internal_value,
                    posto.internal_value,
                    posto.internal_value,
                    posto.internal_value
                )
            )
        prevs.close()
        print('Criado prevs {} de {}'.format(prevs.name, arquivo))

    # Escrita do Vazpast
    vazpast = open(
       file=os.path.join(
           os.path.dirname(arquivo), 'vazpast_{}.dat'.format(os.path.basename(os.path.splitext(arquivo)[0]),)
       ),
       mode='w'
    )

    # Escrita do Export
    export = open(
        file=os.path.join(
            os.path.dirname(arquivo),
            'export_{}.txt'.format(os.path.basename(os.path.splitext(arquivo)[0]),)
        ),
        mode='w'
    )

    # Itera sobre o range completo
    for cols in sheet_vazpast.iter_rows(
            min_row=configs['coords_vazpast']['ini']['row'],
            min_col=configs['coords_vazpast']['ini']['col'],
            max_row=configs['coords_vazpast']['fim']['row'],
            max_col=configs['coords_vazpast']['fim']['col']
    ):
        valores = [i.value for i in cols]
        for j in range(2, 13):
            valores[j] = float(valores[j])

        vazpast.write(
            '{:>5d} ''{:11} ''{:>10.2f}''{:>10.2f}''{:>10.2f}''{:>10.2f}''{:>10.2f}''{:>10.2f}''{:>10.2f}'
            '{:>10.2f}''{:>10.2f}''{:>10.2f}''{:>10.2f}''{:>10.2f}\n'.format(*valores)
        )

        export.write(
            '{:>5d};''{:<15};''{:>10.2f};''{:>10.2f};''{:>10.2f};''{:>10.2f};''{:>10.2f};''{:>10.2f};''{:>10.2f};'
            '{:>10.2f};''{:>10.2f};''{:>10.2f};''{:>10.2f};''{:>10.2f};\n'.format(*valores)
        )

    print('Arquivos vazpast e export do cenario {} criado'.format(os.path.basename(arquivo)))
    work_book.close()

